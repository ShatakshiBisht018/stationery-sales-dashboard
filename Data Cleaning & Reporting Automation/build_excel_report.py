#!/usr/bin/env python3
"""
Generates the automated Excel report (workbook with Cleaned Data,
Cleaning Log, and a formula-driven Summary Dashboard with native charts)
from the output of clean_pipeline.py.
"""
import json
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

FONT = "Arial"
NAVY = "1B3A63"
LIGHT_BLUE = "D9E4F1"
GREEN = "2E7D32"
RED = "C62828"

header_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor=NAVY)
title_font = Font(name=FONT, bold=True, size=16, color=NAVY)
subtitle_font = Font(name=FONT, italic=True, size=10, color="666666")
label_font = Font(name=FONT, bold=True, size=11)
normal_font = Font(name=FONT, size=10)
thin_border = Border(*(Side(style="thin", color="CCCCCC"),) * 4)


def style_header_row(ws, row, n_cols, start_col=1):
    for c in range(start_col, start_col + n_cols):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


def autofit(ws, df, start_col=1):
    for i, col in enumerate(df.columns):
        width = max(12, min(38, int(df[col].astype(str).map(len).max() if len(df) else 10), int(len(str(col)) * 1.3) + 4))
        ws.column_dimensions[get_column_letter(start_col + i)].width = width


def write_df(ws, df, start_row=1, start_col=1, table_name=None, style="TableStyleMedium9"):
    n_rows, n_cols = df.shape
    for j, col in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + j, value=str(col))
    style_header_row(ws, start_row, n_cols, start_col)
    for i in range(n_rows):
        for j, col in enumerate(df.columns):
            val = df.iloc[i, j]
            if pd.isna(val):
                val = ""
            elif hasattr(val, "isoformat"):
                val = val.strftime("%Y-%m-%d")
            cell = ws.cell(row=start_row + 1 + i, column=start_col + j, value=val)
            cell.font = normal_font
            cell.border = thin_border
    autofit(ws, df, start_col)
    if table_name:
        end_col_letter = get_column_letter(start_col + n_cols - 1)
        ref = f"{get_column_letter(start_col)}{start_row}:{end_col_letter}{start_row + n_rows}"
        tbl = Table(displayName=table_name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True)
        ws.add_table(tbl)
    return start_row + n_rows + 1


def build(cleaned_csv, log_json, out_path):
    df = pd.read_csv(cleaned_csv, parse_dates=["Date"])
    with open(log_json) as f:
        log_rows = json.load(f)
    log_df = pd.DataFrame(log_rows)[["timestamp", "action", "detail"]]
    log_df.columns = ["Timestamp", "Action", "Detail"]

    wb = Workbook()

    # ---------------- Sheet 1: Dashboard ----------------
    ws = wb.active
    ws.title = "Dashboard"
    ws["B2"] = "Data Cleaning & Reporting Automation"
    ws["B2"].font = title_font
    ws["B3"] = "Source: stationery_sales_dataset.csv  |  Generated automatically by clean_pipeline.py"
    ws["B3"].font = subtitle_font

    # KPI cells (formulas referencing the Cleaned Data table, not hardcoded)
    kpi_row = 5
    kpis = [
        ("Total Orders", "=COUNTA('Cleaned Data'!A2:A100000)"),
        ("Total Revenue (INR)", "=SUM('Cleaned Data'!K:K)"),
        ("Total Units Sold", "=SUM('Cleaned Data'!H:H)"),
        ("Average Order Value (INR)", "=AVERAGE('Cleaned Data'!K:K)"),
        ("Average Discount (%)", "=AVERAGE('Cleaned Data'!J:J)"),
        ("Outliers Flagged", "=COUNTIF('Cleaned Data'!L:L,\"Yes\")"),
    ]
    for i, (label, formula) in enumerate(kpis):
        r = kpi_row + i
        ws.cell(row=r, column=2, value=label).font = label_font
        cell = ws.cell(row=r, column=4, value=formula)
        cell.font = Font(name=FONT, size=12, bold=True, color=NAVY)
        if "Revenue" in label or "Value" in label:
            cell.number_format = '"₹"#,##0'
        elif "Discount" in label:
            cell.number_format = "0.0\"%\""
        else:
            cell.number_format = "#,##0"

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18

    # Pivot-style table: Revenue by Category (formulas, SUMIF)
    cats = sorted(df["Category"].unique())
    cat_start = kpi_row + len(kpis) + 3
    ws.cell(row=cat_start, column=2, value="Revenue by Category").font = label_font
    ws.cell(row=cat_start + 1, column=2, value="Category").font = header_font
    ws.cell(row=cat_start + 1, column=2).fill = header_fill
    ws.cell(row=cat_start + 1, column=3, value="Revenue (INR)").font = header_font
    ws.cell(row=cat_start + 1, column=3).fill = header_fill
    for i, cat in enumerate(cats):
        r = cat_start + 2 + i
        ws.cell(row=r, column=2, value=cat).font = normal_font
        f = f"=SUMIF('Cleaned Data'!E:E,B{r},'Cleaned Data'!K:K)"
        c = ws.cell(row=r, column=3, value=f)
        c.font = normal_font
        c.number_format = '"₹"#,##0'
    cat_end = cat_start + 1 + len(cats)

    # Pivot-style table: Revenue by Region
    regions = sorted(df["Region"].unique())
    reg_start = cat_end + 3
    ws.cell(row=reg_start, column=2, value="Revenue by Region").font = label_font
    ws.cell(row=reg_start + 1, column=2, value="Region").font = header_font
    ws.cell(row=reg_start + 1, column=2).fill = header_fill
    ws.cell(row=reg_start + 1, column=3, value="Revenue (INR)").font = header_font
    ws.cell(row=reg_start + 1, column=3).fill = header_fill
    for i, reg in enumerate(regions):
        r = reg_start + 2 + i
        ws.cell(row=r, column=2, value=reg).font = normal_font
        f = f"=SUMIF('Cleaned Data'!F:F,B{r},'Cleaned Data'!K:K)"
        c = ws.cell(row=r, column=3, value=f)
        c.font = normal_font
        c.number_format = '"₹"#,##0'
    reg_end = reg_start + 1 + len(regions)

    # Charts
    bar = BarChart()
    bar.title = "Revenue by Category"
    bar.y_axis.title = "Revenue (INR)"
    data = Reference(ws, min_col=3, min_row=cat_start + 1, max_row=cat_end)
    cats_ref = Reference(ws, min_col=2, min_row=cat_start + 2, max_row=cat_end)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.width, bar.height = 15, 9
    ws.add_chart(bar, "F5")

    pie = PieChart()
    pie.title = "Revenue Share by Region"
    data = Reference(ws, min_col=3, min_row=reg_start + 1, max_row=reg_end)
    cats_ref = Reference(ws, min_col=2, min_row=reg_start + 2, max_row=reg_end)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats_ref)
    pie.width, pie.height = 15, 9
    ws.add_chart(pie, "F23")

    # ---------------- Sheet 2: Monthly Trend ----------------
    ws2 = wb.create_sheet("Monthly Trend")
    monthly = df.set_index("Date").resample("ME")["Revenue (INR)"].sum().reset_index()
    monthly["Month"] = monthly["Date"].dt.strftime("%b %Y")
    monthly = monthly[["Month", "Revenue (INR)"]]
    end_row = write_df(ws2, monthly, start_row=2, start_col=2, table_name="MonthlyRevenue")
    ws2.cell(row=1, column=2, value="Monthly Revenue Trend").font = title_font

    line = LineChart()
    line.title = "Monthly Revenue Trend"
    line.y_axis.title = "Revenue (INR)"
    line.x_axis.title = "Month"
    data = Reference(ws2, min_col=3, min_row=2, max_row=1 + len(monthly))
    cats_ref = Reference(ws2, min_col=2, min_row=3, max_row=1 + len(monthly))
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats_ref)
    line.width, line.height = 22, 11
    ws2.add_chart(line, "F2")

    # ---------------- Sheet 3: Cleaned Data ----------------
    ws3 = wb.create_sheet("Cleaned Data")
    write_df(ws3, df, start_row=1, start_col=1, table_name="CleanedData")
    ws3.freeze_panes = "A2"
    # conditional formatting: highlight outliers
    outlier_col = get_column_letter(list(df.columns).index("Outlier Flag") + 1)
    rng = f"{outlier_col}2:{outlier_col}{len(df)+1}"
    ws3.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Yes"'], fill=PatternFill("solid", fgColor="FFF3CD"))
    )

    # ---------------- Sheet 4: Cleaning Log ----------------
    ws4 = wb.create_sheet("Cleaning Log")
    ws4.cell(row=1, column=2, value="Automated Cleaning Log").font = title_font
    ws4.cell(row=2, column=2, value="Every action the pipeline took on this run, in order.").font = subtitle_font
    write_df(ws4, log_df, start_row=4, start_col=2, table_name="CleaningLog")
    ws4.column_dimensions["D"].width = 90

    wb.save(out_path)
    print(f"Saved {out_path}")


if __name__ == "__main__":
    build(sys.argv[1], sys.argv[2], sys.argv[3])
